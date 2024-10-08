import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime

from invoice_s3 import InvoiceS3
from invoice_database import InvoiceDatabase
from invoice import Invoice, CalendarData_Invoice_LinkedList, _copy_range, EXCEL_INVOICE_FILENAME

FILEPATH = r"template/Teacher Base Template.xlsx"
TEACHER_INVOICE_OUTPUT = r"base_generated/teacher/{}/{}.xlsx"

class TeacherInvoice(Invoice):
    def __init__(self, timeMin: datetime, timeMax: datetime, invoice_database: InvoiceDatabase, invoice_s3: InvoiceS3):
        super().__init__(timeMin, timeMax, invoice_database, invoice_s3)

    def write_invoice_to_excel(self) -> None:
        tempHead = self._head
        while tempHead is not None:
            self._total_amount = 0
            revised_date_filename = self._timeMax.strftime(EXCEL_INVOICE_FILENAME).format(
                "A" if self._timeMax.day == 15 else "B", tempHead.person_name)
            destination = (TEACHER_INVOICE_OUTPUT.format(tempHead.person_name, revised_date_filename))

            is_filepath_validated = self._check_filepath_validation(tempHead.person_name, FILEPATH, destination, revised_date_filename)
            if not is_filepath_validated:
                tempHead = tempHead.next
                continue

            # Open workbook and activate current sheet
            local_workbook = openpyxl.load_workbook(destination)
            local_workbook["root"].title = revised_date_filename
            local_sheet = local_workbook.active

            timeMin_str = self._timeMin.strftime("%m/%d")
            timeMax_str = self._timeMax.strftime("%m/%d")
            local_sheet["G3"] = "{}-{}".format(timeMin_str, timeMax_str)

            if local_sheet.cell(row=2, column=7).value == None:
                local_sheet["G2"] = tempHead.invoice_id
            local_sheet["F5"] = tempHead.person_name

            i = 11
            total_hours = 0
            prevtempInvoiceHead = None  # Had to include previous node since rows insert above from the current index.
            tempInvoiceHead = tempHead.invoice
            while tempInvoiceHead != None:
                if local_sheet.cell(row=i, column=5).value != None:
                    i += 1
                    tempInvoiceHead = tempInvoiceHead.next
                    continue
                if local_sheet.cell(row=i+2, column=5).value != None:
                    local_sheet.insert_rows(i+1)
                    _copy_range("B{}:G{}".format(i, i), local_sheet, 1)
                if prevtempInvoiceHead != None:
                    if tempInvoiceHead.person_name.upper() != prevtempInvoiceHead.person_name.upper():
                        local_sheet.insert_rows(i+1)
                        _copy_range("B{}:G{}".format(i, i), local_sheet, 1)
                        i += 1

                if tempInvoiceHead.no_show:
                    bg_color = "bbff99"
                    local_sheet[r"B{}".format(i)].fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                    local_sheet[r"C{}".format(i)].fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                    local_sheet[r"D{}".format(i)].fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                    local_sheet[r"E{}".format(i)].fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                    local_sheet[r"F{}".format(i)].fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                    local_sheet[r"G{}".format(i)].fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")

                local_sheet[r"B{}".format(i)] = tempInvoiceHead.date
                local_sheet[r"C{}".format(i)] = tempInvoiceHead.subject
                local_sheet[r"D{}".format(i)] = tempInvoiceHead.person_name   # Teacher name
                local_sheet[r"E{}".format(i)] = tempInvoiceHead.rate
                local_sheet[r"F{}".format(i)] = tempInvoiceHead.hour
                local_sheet[r"G{}".format(i)] = tempInvoiceHead.amount

                total_hours += tempInvoiceHead.hour
                self._total_amount += tempInvoiceHead.amount

                i += 1
                prevtempInvoiceHead = tempInvoiceHead
                tempInvoiceHead = tempInvoiceHead.next

            for i in range(25, 499):
                if not str(local_sheet.cell(row=i, column=5).value).startswith("Hours"):
                    continue

                local_sheet.cell(row=i, column=6).value = total_hours
                break

            for i in range(31, 499):
                if not str(local_sheet.cell(row=i, column=6).value).startswith("Total"):
                    continue

                local_sheet.cell(row=i, column=7).value = self._total_amount
                break

            # TODO: Convert excel to pdf for safety issue
            local_workbook.save(destination)
            local_workbook.close()

            if not self._invoice_s3 is None:
                self._invoice_s3.upload_file(destination, destination)

            tempHead = tempHead.next
        print("Finished writing teacher excel sheet!")
