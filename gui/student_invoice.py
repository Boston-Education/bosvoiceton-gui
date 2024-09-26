import os
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime

from invoice_s3 import InvoiceS3
from invoice_database import InvoiceDatabase
from invoice import Invoice, CalendarData_Invoice_LinkedList, _copy_range

EXCEL_INVOICE_FILENAME = "%B %Y - ({})"

FILEPATH = r"template/Student Base Template.xlsx"
STUDENT_INVOICE_OUTPUT = r"base_generated/student/{}/{}.xlsx"

def read_past_tuition_amount(destination: str, student_name: str) -> float:
    try:
        previous_path = Path(destination)
        previous_workbook = openpyxl.load_workbook(previous_path)

        previous_sheet = previous_workbook.active
        for i in range(28, 499):
            if not str(previous_sheet.cell(row=i, column=6).value).startswith("Total"):
                continue

            prev_result = previous_sheet.cell(row=i, column=7).value
            previous_workbook.close()
            return prev_result
        return 0.0
    except FileNotFoundError:
        print("ERROR: Unable to find a previous file to access past cost for {}! Reading 0.0 instead...".format(student_name))
        return 0.0

def update_tuition_amount(student_name: str, payment_amount: float, payment_type: str) -> None:
    dest_to_update = Path(r"base_generated/student/{}".format(student_name))
    path_result = [x for x in sorted(dest_to_update.iterdir(), key=os.path.getctime) if x.is_file()]
    #tuition_result = read_past_tuition_amount(path_result[-1].__str__(), path_result[-1].name)

    latest_file_creation = path_result[-1]
    update_workbook = openpyxl.load_workbook(latest_file_creation)
    update_sheet = update_workbook.active
    for i in range(28, 499):
        if not str(update_sheet.cell(row=i, column=6).value).startswith("Total"):
            continue

        current_tuition = update_sheet.cell(row=i, column=7).value
        update_sheet.cell(row=i, column=7).value = current_tuition - payment_amount

        if update_sheet.cell(row=i-2, column=7).value != None or update_sheet.cell(row=i-1, column=7).value != None:
            update_sheet.insert_rows(i - 1)
            _copy_range("B{}:G{}".format(i, i), update_sheet, -1)
        else:
            i -= 1
        update_sheet[r"C{}".format(i)] = datetime.now().strftime("%B %d, %Y")
        update_sheet[r"C{}".format(i)].font = Font(color="009600", size=14)
        update_sheet[r"D{}".format(i)] = payment_type
        update_sheet[r"D{}".format(i)].font = Font(color="009600", size=14)
        update_sheet[r"G{}".format(i)] = -payment_amount
        update_sheet[r"G{}".format(i)].font = Font(color="009600", size=14)

        update_workbook.save(latest_file_creation)
        update_workbook.close()
        return

def _name_previous_excel_date(schedule_date: datetime, student_name: str) -> str:
    # Not to be confused with past invoice date cycle since this algorithm is in reverse.
    month = schedule_date.month - 1
    year = schedule_date.year - 1 if schedule_date.month == 1 else schedule_date.year

    if year < schedule_date.year:
        month = 12

    # Datetime for Day is redundant. Had to insert b/c it's required
    datetime_result = datetime(year, month, 1)
    return datetime_result.strftime(EXCEL_INVOICE_FILENAME.format(student_name))

class StudentInvoice(Invoice):
    def __init__(self, timeMin: datetime, timeMax: datetime, invoice_database: InvoiceDatabase, invoice_s3: InvoiceS3):
        super().__init__(timeMin, timeMax, invoice_database, invoice_s3)

    def write_invoice_to_excel(self) -> None:
        tempHead = self._head
        while tempHead is not None:
            self._total_amount = 0
            revised_date_filename = self._timeMax.strftime(EXCEL_INVOICE_FILENAME).format(tempHead.person_name)
            destination = (STUDENT_INVOICE_OUTPUT.format(tempHead.person_name, revised_date_filename))

            is_filepath_validated = self._check_filepath_validation(tempHead.person_name, FILEPATH, destination, revised_date_filename)
            if not is_filepath_validated:
                tempHead = tempHead.next
                continue

            local_workbook = openpyxl.load_workbook(destination)
            local_workbook["root"].title = revised_date_filename

            local_sheet = local_workbook.active

            prev_excel_date = _name_previous_excel_date(self._timeMax, tempHead.person_name)
            tuition_amount = read_past_tuition_amount(STUDENT_INVOICE_OUTPUT
                                                      .format(tempHead.person_name, prev_excel_date), tempHead.person_name)
            local_sheet["G14"] = tuition_amount
            #self._total_amount += local_sheet["G14"].value
            local_sheet["G3"] = self._today_date.strftime("%b. %d, %Y")
            if local_sheet.cell(row=4, column=7).value == None:
                local_sheet["G4"] = tempHead.invoice_id
            local_sheet["G6"] = tempHead.person_name

            i = 16
            prevtempInvoiceHead = None  # Had to include previous node since rows insert above from the current index.
            tempInvoiceHead = tempHead.invoice
            while tempInvoiceHead != None:
                if local_sheet.cell(row=i, column=2).value != None:
                    i += 1
                    tempInvoiceHead = tempInvoiceHead.next
                    continue
                if local_sheet.cell(row=i+2, column=2).value != None:
                    local_sheet.insert_rows(i+1)
                    _copy_range("B{}:G{}".format(i, i), local_sheet, 1)
                if prevtempInvoiceHead != None:
                    if tempInvoiceHead.person_name.title() != prevtempInvoiceHead.person_name.title():
                        local_sheet.insert_rows(i+1)
                        _copy_range("B{}:G{}".format(i, i), local_sheet, 1)
                        i += 1

                if tempInvoiceHead.no_show:
                    bg_color = "bbff99"
                    local_sheet[r"B{}".format(i)].fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                    local_sheet[r"C{}".format(i)].fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                    local_sheet[r"D{}".format(i)].fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                    local_sheet[r"E{}".format(i)].fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                    local_sheet[r"F{}".format(i)].fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                    local_sheet[r"G{}".format(i)].fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")

                if tempInvoiceHead.amount <= 0:
                    font_color = "ff0000"
                    local_sheet[r"B{}".format(i)].font = Font(color=font_color, size=14)
                    local_sheet[r"C{}".format(i)].font = Font(color=font_color, size=14)
                    local_sheet[r"D{}".format(i)].font = Font(color=font_color, size=14)
                    local_sheet[r"E{}".format(i)].font = Font(color=font_color, size=14)
                    local_sheet[r"F{}".format(i)].font = Font(color=font_color, size=14)
                    local_sheet[r"G{}".format(i)].font = Font(color=font_color, size=14)

                local_sheet[r"B{}".format(i)] = tempInvoiceHead.date
                local_sheet[r"C{}".format(i)] = tempInvoiceHead.subject
                local_sheet[r"D{}".format(i)] = tempInvoiceHead.person_name   # Teacher name
                local_sheet[r"E{}".format(i)] = tempInvoiceHead.rate
                local_sheet[r"F{}".format(i)] = tempInvoiceHead.hour
                local_sheet[r"G{}".format(i)] = tempInvoiceHead.amount

                self._total_amount += tempInvoiceHead.amount

                i += 1
                prevtempInvoiceHead = tempInvoiceHead
                tempInvoiceHead = tempInvoiceHead.next

            discount_amount = self._invoice_database.get_discount_amount(tempHead.person_name)
            if discount_amount != None:
                if local_sheet.cell(row=i+2, column=2).value != None:
                    local_sheet.insert_rows(i+1)
                    _copy_range("B{}:G{}".format(i, i), local_sheet, 1)
                    i += 1
                else:
                    i += 1
                local_sheet[r"D{}".format(i)] = "Discount:"
                local_sheet[r"D{}".format(i)].font = Font(color="FF0000", size=14)
                local_sheet[r"E{}".format(i)] = str(discount_amount * 100) + "%"
                local_sheet[r"E{}".format(i)].font = Font(color="FF0000", size=14)
                local_sheet[r"G{}".format(i)] = (self._total_amount * (1 - discount_amount))
                local_sheet[r"G{}".format(i)].font = Font(color="FF0000", size=14)

            for i in range(28, 499):
                if not str(local_sheet.cell(row=i, column=6).value).startswith("Total"):
                    continue

                local_sheet.cell(row=i, column=7).value = self._total_amount + tuition_amount

                if discount_amount == None:
                    break
                if discount_amount != 0:
                    print("Discount applied to student, {}, for {}%!".format(tempHead.person_name, discount_amount * 100))

                discount_cost = local_sheet.cell(row=i, column=7).value = (self._total_amount * (1 - discount_amount)) + tuition_amount
                break

            # TODO: Convert excel to pdf for safety issue
            local_workbook.save(destination)
            local_workbook.close()

            if not self._invoice_s3 is None:
                self._invoice_s3.upload_file(destination, destination)

            tempHead = tempHead.next
        print("Finished writing student excel sheet!")
